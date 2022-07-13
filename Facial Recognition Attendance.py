import sys, os, cv2, shutil, os.path, time, mysql.connector, imutils, requests, math,netifaces
import numpy as np
from threading import Thread
from tkinter import *
from datetime import date
from datetime import datetime
from PIL import Image
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import (QDialog, QApplication, QWidget,
                             QVBoxLayout, QListWidget, QListWidgetItem,QLineEdit)
from PyQt5.QtCore import pyqtSignal, pyqtSlot, Qt, QThread,QTimer
from PyQt5.QtGui import QPixmap, QFont,QIcon
from openpyxl import load_workbook

recognizer = cv2.face.LBPHFaceRecognizer_create()
detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
url = "http://"+str(netifaces.gateways()['default'][netifaces.AF_INET][0])+":8080/shot.jpg"


class StartApp(QDialog):
    def __init__(self):
        loginscreen = LoginScreen()
        widget.addWidget(loginscreen)
        homescreen = HomeScreen()
        widget.addWidget(homescreen)
        self.viewattendance = ViewAttendance()
        widget.addWidget(self.viewattendance)
        self.registerface = RegisterFaces()
        widget.addWidget(self.registerface)
        self.takeattendance = TakeAttendance()
        widget.addWidget(self.takeattendance)
        self.rgstrcourse = RegisterCourses()
        widget.addWidget(self.rgstrcourse)
        self.addinfotakeattendance = AddInfoTakeAttendance()
        widget.addWidget(self.addinfotakeattendance)
        widget.setFixedWidth(1000)
        widget.setFixedHeight(600)
        widget.setCurrentIndex(0)
        widget.show()


class LoginScreen(QDialog):
    def __init__(self):
        super(LoginScreen, self).__init__()
        loadUi("uis/login.ui", self)
        self.loginButton.clicked.connect(self.gotohomescreen)
        self.showps.clicked.connect(self.showPassword)
        self.empty.hide()
        self.wrong1.hide()
        self.wrong2.hide()



    def showPassword(self):
        if self.showps.isChecked():
            self.password.setEchoMode(QLineEdit.Normal)
        else:
            self.password.setEchoMode(QLineEdit.Password)

    def gotohomescreen(self):
        lst = []
        pn = self.username.text()
        ps = self.password.text()
        db = mysql.connector.connect(host="localhost", user="jonpol", password="root", database='USERS')
        cursor = db.cursor()
        cursor.execute("select exists(select 1 from users) as OUTPUT")
        for itm in cursor:
            if itm[0] == 0:
                self.empty.show()
                QTimer.singleShot(2000, self.empty.hide)
            else:
                cursor.reset()
                cursor.execute("select USER_NAME,PASSWORDS from users")
                for itm in cursor:
                    lst.append(itm[0] + '.' + itm[1])
                    
                count1 = 0
                count2 = 0
                count3 = 0    
                for tm in lst:
                    if pn == tm.split('.')[0] and ps == tm.split('.')[1]:
                        widget.setCurrentIndex(1)
                        self.password.setText('')
                        break

                    elif pn != tm.split('.')[0] and ps == tm.split('.')[1]:
                        self.wrong1.show()
                        QTimer.singleShot(2000, self.wrong1.hide)
                        count1 += 1

                    elif pn == tm.split('.')[0] and ps != tm.split('.')[1]:
                        count2 += 1

                    elif pn != tm.split('.')[0] and ps != tm.split('.')[1]:
                        count3 += 1

                if count1 > 0:
                    self.wrong1.show()
                    QTimer.singleShot(2000, self.wrong1.hide)
                elif count2 > 0:
                    self.wrong2.show()
                    QTimer.singleShot(2000, self.wrong2.hide)
                elif count3 > 0:
                    self.wrong1.show()
                    self.wrong2.show()
                    QTimer.singleShot(2000, self.wrong1.hide)
                    QTimer.singleShot(2000, self.wrong2.hide)


class HomeScreen(QDialog):
    def __init__(self):
        super(HomeScreen, self).__init__()
        loadUi("uis/homescreen.ui", self)
        self.takeattendanceButton.clicked.connect(self.gotoaddinfotakeattendance)
        self.viewattendanceButton.clicked.connect(self.gotoviewattendance)
        self.registerfaceButton.clicked.connect(self.gotoregisterface)
        self.registercoursesButton.clicked.connect(self.gotoregistercourses)
        self.lg.clicked.connect(self.gotologin)


    def gotoaddinfotakeattendance(self):
        for item in range(rgc.comboBox4.count()):
            ad.comboBox.addItem(rgc.comboBox4.itemText(item))
        widget.setCurrentIndex(6)

    def gotoviewattendance(self):
        for item in range(rgc.comboBox4.count()):
            vat.comboBox.addItem(rgc.comboBox4.itemText(item))
        widget.setCurrentIndex(2)

    def gotoregisterface(self):
        widget.setCurrentIndex(3)

    def gotoregistercourses(self):
        widget.setCurrentIndex(5)

    def gotologin(self):
        widget.setCurrentIndex(0)


class TakeAttendance(QDialog):
    def __init__(self):
        super(TakeAttendance, self).__init__()
        loadUi("uis/takeattendancescreen.ui", self)
        self.status = False
        self.saved.hide()
        self.alreadysaved.hide()
        self.stopped.hide()
        self.nctc.hide()
        self.startButton.clicked.connect(self.starter)
        self.gobackButton.clicked.connect(self.goback)
        self.saveButton.clicked.connect(self.save)
        self.stopButton.clicked.connect(self.stoper)
        self.init = 0

    def starter(self):
        open("collector.txt", "w")
        open("collected.txt", "w")
        widget = QWidget()
        vbox = QVBoxLayout()
        widget.setLayout(vbox)
        self.scroll.setWidget(widget)
        ls = QListWidget()
        QListWidgetItem("", ls)
        ls.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
        vbox.addWidget(ls)
        self.nctc.hide()
        self.init = time.time()
        ad.vdstate = True
        self.status = False
        # create the video capture thread
        self.thread = VideoThread1()
        # connect its signal to the update_image slot
        self.thread.change_pixmap_signal.connect(self.thread.update_image)
        # start the thread
        self.t = Thread(target=self.thread.run)
        self.t.start()

    def stoper(self):
        ad.vdstate = False
        self.status = True
        time.sleep(1)
        self.t.join()
        self.stopped.show()
        QTimer.singleShot(2000,self.stopped.hide)

    def save(self):
        dt = str(ad.date.date().toPyDate())
        ndt = dt.split('-')
        book = load_workbook('original2.xlsx')
        sheet = book.active
        com = ad.comboBox.currentText().split(':')
        pth = "sheets"
        sheet["A1"] = "UNIVERSITY OF DAR ES SALAAM\n" \
                      "COLLEGE OF INFORMATION AND COMMUNICATION TECHNOLOGIES\n" \
                      "(CoICT)\n" \
                      "DEPARTMENT OF ELECTRONICS AND TELECOMMUNICATIONS\n" + \
                      ad.comboBox.currentText() + "\n" \
                                                  "ATTENDANCE SHEET"
        n = 1
        a = open("collected.txt", "r")
        for lines in a:
            ln = lines.split('.')
            tm = int(ln[3].split()[0]) - self.init
            sheet["A" + str(n + 9)] = n
            sheet["B" + str(n + 9)] = ln[0]
            sheet["C" + str(n + 9)] = ln[1].split()[0]
            sheet["D" + str(n + 9)] = ln[2].split()[0]
            sheet["E" + str(n + 9)] = str(round(tm / 60)) + ":" + str(round(tm % 60))
            n += 1
        cm = com[0].split()[0]
        dir = [dr for dr in os.listdir(pth)]
        for n in range(ad.comboBox.count()):
            dr1 = dir[n].split()[0]
            if cm == dr1:
                if os.path.isfile("sheets/" + dr1 + "/" + ndt[2] + ndt[1] + ".xlsx") == False:
                    book.save(filename="sheets/" + dr1 + "/" + ndt[2] + ndt[1] + ".xlsx")
                    book1 = load_workbook("sheets/" + dr1 + "/" + dr1 + ".xlsx")
                    book2 = load_workbook("sheets/" + dr1 + "/" + ndt[2] + ndt[1] + ".xlsx")
                    sheet1 = book1.active
                    sheet2 = book2.active
                    storer = []
                    for row in range(10, sheet1.max_row + 1):
                        if sheet1.cell(row=row, column=1).value != None:
                            mux_row = sheet1.cell(row=row, column=1).value
                    columncount = 5
                    for column in range(5, sheet1.max_column):
                        if sheet1.cell(row=9, column=column).value != None:
                            columncount = column + 1
                    sheet1.cell(row=9, column=columncount, value=ndt[2] + "/" + ndt[1])
                    for x in range(10, mux_row + 10):
                        count = 0
                        cell_obj1 = sheet1.cell(row=x, column=3)
                        for y in range(10, mux_row + 10):
                            cell_obj2 = sheet2.cell(row=y, column=3)
                            if cell_obj1.value == cell_obj2.value:
                                count += 1
                        if count > 0:
                            sheet1.cell(row=x, column=columncount, value=u"\u2713")
                        else:
                            sheet1.cell(row=x, column=columncount, value="x")
                        counter = []
                        cnt = 0
                        for m in range(5, columncount + 1):
                            if sheet1.cell(row=x, column=m).value == u"\u2713":
                                counter.append(1)
                            else:
                                counter.append(0)
                            cnt += counter[m - 5]
                        storer.append(cnt)
                        sheet1.cell(row=x, column=4, value=int((storer[x - 10] / (columncount - 4)) * 100))
                    book1.save(filename="sheets/" + dr1 + ".xlsx")
                    os.remove("sheets/" + dr1 + "/" + dr1 + ".xlsx")
                    shutil.move("sheets/" + dr1 + ".xlsx", "sheets/" + dr1)
                    self.saved.show()
                    QTimer.singleShot(2000,self.saved.hide)

                else:
                    self.alreadysaved.show()
                    QTimer.singleShot(2000,self.alreadysaved.hide)

                    break

    def goback(self):
        widget.setCurrentIndex(6)


class VideoThread1(QThread):
    @pyqtSlot(np.ndarray)
    def update_image(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        faces = detector.detectMultiScale(rgb_image, scaleFactor=1.3, minNeighbors=5, )
        for (x, y, w, h) in faces:
            cv2.rectangle(cv_img, (x, y), (x + w, y + h), (255, 255, 255), 2)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(399, 349, Qt.KeepAspectRatio)
        """Updates the image_label with a new opencv image"""
        qt_img = QPixmap.fromImage(p)
        st.label.setPixmap(qt_img)
        a = open('collector.txt', 'r')
        b = open('collected.txt', 'r')
        if (len(a.readlines()) != len(b.readlines())):
            widget = QWidget()
            vbox = QVBoxLayout()
            widget.setLayout(vbox)
            st.scroll.setWidget(widget)
            ls = QListWidget()
            a = open('collector.txt', 'r')
            open('collected.txt', 'w+')
            n = 1
            for lines in a:
                b = open('collected.txt', 'a+')
                b.write(lines)
                QListWidgetItem(str(n) + "." + str(lines.split('.')[0]) + "\n  " +
                                str(lines.split('.')[1]).ljust(55) + str(lines.split('.')[2]) + "\n\n", ls)
                ls.setFont(QFont('Arial', 13))
                ls.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
                n += 1
            vbox.addWidget(ls)

    change_pixmap_signal = pyqtSignal(np.ndarray)

    def run(self):
        id = []
        names = ['unknown']
        reg = ['unknown']
        pth = os.listdir("faces/" + ad.comboBox.currentText().split()[0])
        for m in range(1, int(pth[-1].split('.')[0]) + 1):
            names.append(pth[(m - 1) * 30].split('.')[2])
            reg.append(pth[(m - 1) * 30].split('.')[3])
            id.append(m)
        # capture from web cam
        try:
            while True:
                img_resp = requests.get(url)
                img_arr = np.array(bytearray(img_resp.content), dtype=np.uint8)
                img = cv2.imdecode(img_arr, -1)
                imgs = imutils.resize(img, width=1920, height=1080)
                cv_img = cv2.flip(imgs, 1)

                gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
                faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, )
                for (x, y, w, h) in faces:
                    
                    height, width, channels = cv_img.shape
                    cropped = cv_img[y-80:y + h + 80, x-160:x + w + 160]
                    cv_img = cv2.resize(cropped, (width, height))
                    
                    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
                    faces = detector.detectMultiScale(gray, 1.3, 5)
                    for (x, y, w, h) in faces:
                        cv2.rectangle(cv_img, (x, y), (x + w, y + h), (255, 255, 255), 2)
                        id, confidence = recognizer.predict(gray[y:y + h, x:x + w])
                        # Check if confidence is less them 100 ==> "0" is perfect match
                        if (confidence < 100-(int(st.pcnt.currentText()))):
                            ida = names[id]
                            rg = reg[id]
                            confidence = "  {0}%".format(round(100 - confidence))
                            f = open('collector.txt', 'a+')
                            g = open('collector.txt', 'r')
                            flag = 0
                            ind = 0
                            for line in g:
                                ind += 1
                                if ida in line:
                                    flag = 1
                                    break
                            if (flag == 0 and id > 0):
                                now = datetime.now()
                                current_time = now.strftime("%H:%M:%S")
                                f.write(ida + '.' + rg + '.' + current_time + '.' + str(time.time()) + '\n')
                        else:
                            ida = 'unknown'
                            confidence = "  {0}%".format(round(100 - confidence))
                        cv2.putText(cv_img, str(confidence), (5,30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255),2)

                if True:
                    self.change_pixmap_signal.emit(cv_img)

                if st.status == True:
                    break
        except:
            st.nctc.show()
            QTimer.singleShot(2000,st.nctc.hide)



class RegisterCourses(QDialog):
    def __init__(self):
        super(RegisterCourses, self).__init__()
        loadUi("uis/registercoursescreen.ui", self)
        self.comment.hide()
        self.comment1.hide()
        self.addButton.clicked.connect(self.additems)
        self.removeButton.clicked.connect(self.removeitems)
        self.clearallButton.clicked.connect(self.clearallitems)
        self.cancelButton.clicked.connect(self.goback)
        self.saveButton.clicked.connect(self.modifycoursesfile)
        self.saveButton.clicked.connect(self.createfilesandsheets)
        self.comboBox1.activated.connect(self.addyears)
        self.comboBox2.activated.connect(self.addcourses)

    def addyears(self):
        db = mysql.connector.connect(host="localhost", user="jonpol", password="root", database='courses')
        cursor = db.cursor()
        self.comboBox2.clear()
        self.comboBox3.clear()
        cursor.execute("select " + self.comboBox1.currentText() + " from FUCULTIES")
        for course in cursor:
            self.comboBox2.addItem(course[0])

    def addcourses(self):
        db = mysql.connector.connect(host="localhost", user="jonpol", password="root", database='courses')
        cursor = db.cursor()
        self.comboBox3.clear()
        cursor = db.cursor()
        cursor.execute("select " + self.comboBox2.currentText() + " from TE_COURSES")
        for course in cursor:
            self.comboBox3.addItem(course[0])

    def additems(self):
        counter = 0
        if self.comboBox3.count() != 0:
            for i in range(self.comboBox4.count()):
                item1 = self.comboBox4.itemText(i)
                item2 = self.comboBox3.currentText()
                if (item1 == item2):
                    counter += 1
            if (counter == 1):
                self.comment.show()
                QTimer.singleShot(2000,self.comment.hide)

            else:
                self.comboBox4.addItem(self.comboBox3.currentText())
                self.comboBox4.setCurrentIndex(self.comboBox4.count() - 1)

    def removeitems(self):
        self.comboBox4.removeItem(self.comboBox4.currentIndex())
        
    def clearallitems(self):
        self.comboBox4.clear()
        

    def goback(self):
        widget.setCurrentIndex(1)
        

    def modifycoursesfile(self):
        savedlogs = open("savinglogs.log")
        num1 = self.comboBox4.count()
        num2 = len(savedlogs.readlines())
        if (5 * num1 != num2):
            filo = open("uis/courses.ui")

            middle = open("middlepart.txt", "a+")
            for j in range(self.comboBox4.count()):
                middle.write(' <item>\n\t <property name="text">\n\t  <string>' + self.comboBox4.itemText(
                    j) + '</string>\n\t </property>\n </item>\n')

            first = open("firstpart.txt", "a+")
            for i in range(219):
                lines1 = filo.readline()
                first.write(lines1)

            last = open("lastpart.txt", "a+")
            for i in range(220, 391):
                lines2 = filo.readline()
                last.write(lines2)

            first = open("firstpart.txt")
            middle = open("middlepart.txt")
            last = open("lastpart.txt")
            file = open("uis/registercoursescreen.ui", "w+")

            for i in range(219):
                lines_a = first.readline()
                file.write(lines_a)

            for i in range(5 * self.comboBox4.count()):
                lines_b = middle.readline()
                file.write(lines_b)

            for i in range(171):
                lines_c = last.readline()
                file.write(lines_c)

            first.close()
            middle.close()
            last.close()
            file.close()
            savelogs = open("savinglogs.log", "w+")
            savelogs.write(str(5 * self.comboBox4.count()))
            os.remove("firstpart.txt")
            os.remove("middlepart.txt")
            os.remove("lastpart.txt")
            self.comment1.show()
            QTimer.singleShot(2000,self.comment1.hide)
            addinfotakeattendance = AddInfoTakeAttendance()
            widget.insertWidget(7, addinfotakeattendance)

    def createfilesandsheets(self):
        path = "sheets"
        dir = os.listdir(path)
        for item in range(self.comboBox4.count()):
            count = 0
            counnt = 0
            itm = self.comboBox4.itemText(item).split(':')
            for dr in dir:
                if itm[0].split()[0] == dr:
                    count += 1
                    counnt += 1
            if count < 1:
                os.makedirs("sheets/" + itm[0].split()[0])
                os.makedirs("faces/" + itm[0].split()[0])
                os.makedirs("reports/" + itm[0].split()[0])
                book = load_workbook('original1.xlsx')
                book1 = load_workbook('original3.xlsx')
                sheet = book.active
                sheet1 =book1.active
                sheet["A1"] = "UNIVERSITY OF DAR ES SALAAM\n" \
                              "COLLEGE OF INFORMATION AND COMMUNICATION TECHNOLOGIES\n" \
                              "(CoICT)\n" \
                              "DEPARTMENT OF ELECTRONICS AND TELECOMMUNICATIONS\n" \
                              + self.comboBox4.itemText(item) + "\n" \
                                                                "ATTENDANCE SHEET"
                book.save(filename=("sheets/"+itm[0].split()[0]+".xlsx"))
                shutil.move("sheets/"+itm[0].split()[0] + ".xlsx", "sheets/" + itm[0].split()[0])

                
                sheet1["A1"] = "UNIVERSITY OF DAR ES SALAAM\n" \
                              "COLLEGE OF INFORMATION AND COMMUNICATION TECHNOLOGIES\n" \
                              "(CoICT)\n" \
                              "DEPARTMENT OF ELECTRONICS AND TELECOMMUNICATIONS\n" \
                              + self.comboBox4.itemText(item) + "\n" \
                                                                "REPORT SHEET"
                fl= "reports/" + itm[0].split()[0]
                book1.save(filename=("reports/" + itm[0].split()[0]+".xlsx"))
                shutil.move("reports/" + itm[0].split()[0]+".xlsx", "reports/" + itm[0].split()[0]+"/"+itm[0].split()[0]+".xlsx")

        for dr in dir:
            count = 0
            for itm in range(self.comboBox4.count()):
                if dr == self.comboBox4.itemText(itm).split()[0]:
                    count += 1
            if count < 1:
                shutil.rmtree("faces/" + dr)
                shutil.rmtree("sheets/" + dr)
                shutil.rmtree("reports/" + dr)


class ViewAttendance(QDialog):
    def __init__(self):
        super(ViewAttendance, self).__init__()
        loadUi("uis/viewattendancescreen.ui", self)
        self.ndt.hide()
        self.nofile.hide()
        self.dn.hide()
        date_str = str(date.today())
        # convert str to QDate
        qdate = QtCore.QDate.fromString(date_str, "yyyy-MM-d")
        # Set the format of how the QDate will be displayed in the widget
        self.date.setDisplayFormat("d-MMM-yyyy")
        self.date.setDate(qdate)
        self.gobackButton2.clicked.connect(self.goback)
        self.chekstatsButton.clicked.connect(self.chekstats)
        self.generateButton.clicked.connect(self.generate)
        self.openatt.clicked.connect(self.openat)
        self.openexct.clicked.connect(self.openatex)

    def chekstats(self):
        self.nofile.hide()
        self.ndt.hide()
        if self.comboBox.count() != 0:
            com = self.comboBox.currentText().split(':')
            cm = com[0].split()[0]
            pth = "sheets"
            path = "sheets/" + cm
            dir = [dr for dr in os.listdir(pth)]
            ltcm = []
            files = os.listdir(path)
            if len(files) > 1:
                widget = QWidget()
                vbox = QVBoxLayout()
                widget.setLayout(vbox)
                self.scroll1.setWidget(widget)
                ls = QListWidget()
                for sht in range(len(files) - 1):
                    book = load_workbook(path + "/" + files[sht])
                    sheet = book.active
                    for i in range(10, sheet.max_row + 1):
                        if sheet.cell(row=i, column=5).value != None:
                            vlv = sheet.cell(row=i, column=5).value.split(':')
                            if int(vlv[0]) >= 30:
                                ltcm.append(str(sheet.cell(row=i, column=2).value)
                                            + "." + str(sheet.cell(row=i, column=3).value))
                        else:
                            break

                last = []
                for j in range(len(ltcm)):
                    count = 0
                    for k in range(len(ltcm)):
                        if ltcm[j] == ltcm[k]:
                            count += 1
                    last.append(str(ltcm[j]) + "." + str(count))
                lst = list(dict.fromkeys(last))
                book1 = load_workbook(path + "/" + cm + ".xlsx")
                sht1 = book1.active
                nlst = []
                for i in range(10, sht1.max_row + 1):
                    if sht1.cell(row=i, column=2).value != None:
                        lvll = sht1.cell(row=i, column=4).value
                        lvll = math.ceil(lvll * (len(files) - 1) / 100)
                        for j in lst:
                            if sht1.cell(row=i, column=2).value == j.split('.')[0]:
                                if int(j.split('.')[-1]) >= math.ceil(lvll / 2):
                                    nlst.append(j + "." + str(lvll))

                nlst.sort()
                n = 1
                for lt in nlst:
                    QListWidgetItem(str(n) + "." + lt.split(".")[0] + "\n   " + lt.split(".")[1]
                                    + "\t" + lt.split(".")[-2] + " Out of " + lt.split(".")[-1] + " Days" + "\n", ls)
                    ls.setFont(QFont('Arial', 13))
                    ls.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
                    vbox.addWidget(ls)
                    n += 1

                for n in range(self.comboBox.count()):
                    dr1 = dir[n].split()[0]
                    if cm == dr1:
                        if os.path.isfile("sheets/" + dr1 + "/" + dr1 + ".xlsx") == True:
                            book = load_workbook("sheets/" + dr1 + "/" + dr1 + ".xlsx")
                            sheet = book.active
                            num = []
                            for vl in range(10, sheet.max_row + 1):
                                if sheet.cell(row=vl, column=4).value != None:
                                    num.append(sheet.cell(row=vl, column=4).value)
                                counnt1 = 0
                                counnt2 = 0
                                arng = []
                                last = []
                                for i in range(len(num)):
                                    if num[i] >= 75:
                                        counnt1 += 1
                                    else:
                                        counnt2 += 1
                                    for j in range(len(num)):
                                        cnt = 0
                                        for tm in arng:
                                            if num[i] == tm:
                                                cnt += 1
                                        if cnt == 0:
                                            arng.append(num[i])
                                for i in range(len(arng)):
                                    count = 0
                                    for j in range(len(arng)):
                                        if arng[i] > arng[j]:
                                            count += 1
                                    if count < 3 and arng[i] < 50:
                                        last.append(arng[i])

                if counnt1 == 1:
                    self.textBrowser1.setText(str(counnt1) + "  STUDENT")
                else:
                    self.textBrowser1.setText(str(counnt1) + "  STUDENTS")
                if counnt2 == 1:
                    self.textBrowser2.setText(str(counnt2) + "  STUDENT")
                else:
                    self.textBrowser2.setText(str(counnt2) + "  STUDENTS")
                widget = QWidget()
                vbox = QVBoxLayout()
                widget.setLayout(vbox)
                self.scroll.setWidget(widget)
                ls = QListWidget()
                n = 1
                ltt = []
                for vl in range(10, sheet.max_row + 1):
                    for nm in range(len(last)):
                        if sheet.cell(row=vl, column=4).value == last[nm]:
                            ltt.append(str(sheet.cell(row=vl, column=2).value) + "."
                                       + str(sheet.cell(row=vl, column=3).value) + "."
                                       + str(sheet.cell(row=vl, column=4).value))
                ltt.sort()
                for vl in ltt:
                    QListWidgetItem((str(n) + "." + vl.split(".")[0] + "\n   " + vl.split(".")[1] +
                                     "\t" + vl.split(".")[2] + "%" + "\n"), ls)
                    ls.setFont(QFont('Arial', 13))
                    ls.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
                    vbox.addWidget(ls)
                    n += 1

            else:
                self.textBrowser1.setText("")
                self.textBrowser2.setText("")
                widget1 = QWidget()
                widget2 = QWidget()
                vbox1 = QVBoxLayout()
                vbox2 = QVBoxLayout()
                widget1.setLayout(vbox1)
                widget2.setLayout(vbox2)
                self.scroll.setWidget(widget1)
                self.scroll1.setWidget(widget2)
                ls1 = QListWidget()
                ls2 = QListWidget()
                QListWidgetItem("", ls1)
                QListWidgetItem("", ls2)
                ls1.setFont(QFont('Arial', 13))
                ls2.setFont(QFont('Arial', 13))
                ls1.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
                ls2.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
                vbox1.addWidget(ls1)
                vbox2.addWidget(ls2)
                self.ndt.show()
                QTimer.singleShot(2000,self.ndt.hide)


        else:
            self.textBrowser1.setText("")
            self.textBrowser2.setText("")
            widget1 = QWidget()
            widget2 = QWidget()
            vbox1 = QVBoxLayout()
            vbox2 = QVBoxLayout()
            widget1.setLayout(vbox1)
            widget2.setLayout(vbox2)
            self.scroll.setWidget(widget1)
            self.scroll1.setWidget(widget2)
            ls1 = QListWidget()
            ls2 = QListWidget()
            QListWidgetItem("", ls1)
            QListWidgetItem("", ls2)
            ls1.setFont(QFont('Arial', 13))
            ls2.setFont(QFont('Arial', 13))
            ls1.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
            ls2.setStyleSheet("QLabel""{""border : 1px solid black;""background : white;""}")
            vbox1.addWidget(ls1)
            vbox2.addWidget(ls2)
            self.ndt.show()
            QTimer.singleShot(2000,self.ndt.hide)


    def openat(self):
        if rgc.comboBox4.count() != 0:
            com = self.comboBox.currentText().split(':')
            pth = "sheets"
            cm = com[0].split()[0]
            dir = [dr for dr in os.listdir(pth)]
            for n in range(self.comboBox.count()):
                dr1 = dir[n].split()[0]
                if cm == dr1:
                    if os.path.isfile("sheets/" + dr1 + "/" + dr1 + ".xlsx") == True:
                        os.system("start" + " sheets/" + dr1 + "/" + dr1 + ".xlsx")
        else:
            self.nofile.show()
            QTimer.singleShot(2000,self.nofile.hide)


    def openatex(self):
        if rgc.comboBox4.count() != 0:
            dt = str(self.date.date().toPyDate())
            ndt = dt.split('-')
            com = self.comboBox.currentText().split(':')
            pth = "sheets"
            cm = com[0].split()[0]
            dir = [dr for dr in os.listdir(pth)]
            for n in range(self.comboBox.count()):
                dr1 = dir[n].split()[0]
                if cm == dr1:
                    if os.path.isfile("sheets/" + dr1 + "/" + ndt[2] + ndt[1] + ".xlsx") == True:
                        self.nofile.hide()
                        os.system("start" + " sheets/" + dr1 + "/" + ndt[2] + ndt[1] + ".xlsx")
                    else:
                        self.nofile.show()
                        QTimer.singleShot(2000,self.nofile.hide)

        else:
            self.nofile.show()
            QTimer.singleShot(2000,self.nofile.hide)
            

    def generate(self):
        colD = []
        colE = []

        if self.comboBox.count() != 0:
            com = self.comboBox.currentText().split(':')
            cm = com[0].split()[0]
            pth = "sheets"
            path = "sheets/" + cm
            dir = [dr for dr in os.listdir(pth)]
            ltcm = []
            files = os.listdir(path)
            if len(files) > 1:
                for sht in range(len(files) - 1):
                    book = load_workbook(path + "/" + files[sht])
                    sheet = book.active
                    for i in range(10, sheet.max_row + 1):
                        if sheet.cell(row=i, column=5).value != None:
                            vlv = sheet.cell(row=i, column=5).value.split(':')
                            if int(vlv[0]) >= 30:
                                ltcm.append(str(sheet.cell(row=i, column=2).value)
                                            + "." + str(sheet.cell(row=i, column=3).value))
                        else:
                            break

                last = []
                for j in range(len(ltcm)):
                    count = 0
                    for k in range(len(ltcm)):
                        if ltcm[j] == ltcm[k]:
                            count += 1
                    last.append(str(ltcm[j]) + "." + str(count))
                lst = list(dict.fromkeys(last))
                book1 = load_workbook(path + "/" + cm + ".xlsx")
                sht1 = book1.active
                nlst = []
                for i in range(10, sht1.max_row + 1):
                    if sht1.cell(row=i, column=2).value != None:
                        lvll = sht1.cell(row=i, column=4).value
                        lvll = math.ceil(lvll * (len(files) - 1) / 100)
                        for j in lst:
                            if sht1.cell(row=i, column=2).value == j.split('.')[0]:
                                if int(j.split('.')[-1]) >= math.ceil(lvll / 2):
                                    nlst.append(j + "." + str(lvll))

                nlst.sort()
                for lt in nlst:
                    colE.append(lt.split(".")[0])

                for n in range(self.comboBox.count()):
                    dr1 = dir[n].split()[0]
                    if cm == dr1:
                        if os.path.isfile("sheets/" + dr1 + "/" + dr1 + ".xlsx") == True:
                            book = load_workbook("sheets/" + dr1 + "/" + dr1 + ".xlsx")
                            sheet = book.active
                            num = []
                            for vl in range(10, sheet.max_row + 1):
                                if sheet.cell(row=vl, column=4).value != None:
                                    num.append(sheet.cell(row=vl, column=4).value)
                                counnt1 = 0
                                counnt2 = 0
                                arng = []
                                last = []
                                colB = []
                                colC = []
                                for i in range(len(num)):
                                    if num[i] >= 75:
                                        colB.append(sheet.cell(row=i+10, column=2).value)
                                        counnt1 += 1
                                    else:
                                        colC.append(sheet.cell(row=i+10, column=2).value)
                                        counnt2 += 1
                                    for j in range(len(num)):
                                        cnt = 0
                                        for tm in arng:
                                            if num[i] == tm:
                                                cnt += 1
                                        if cnt == 0:
                                            arng.append(num[i])
                                for i in range(len(arng)):
                                    count = 0
                                    for j in range(len(arng)):
                                        if arng[i] > arng[j]:
                                            count += 1
                                    if count < 3 and arng[i] < 50:
                                        last.append(arng[i])

                ltt = []
                for vl in range(10, sheet.max_row + 1):
                    for nm in range(len(last)):
                        if sheet.cell(row=vl, column=4).value == last[nm]:
                            ltt.append(str(sheet.cell(row=vl, column=2).value) + "."
                                       + str(sheet.cell(row=vl, column=3).value) + "."
                                       + str(sheet.cell(row=vl, column=4).value))
                ltt.sort()
                for vl in ltt:
                    colD.append(vl.split(".")[0])
                self.dn.show()
                QTimer.singleShot(2000,self.dn.hide)

            else:
                self.ndt.show()
                QTimer.singleShot(2000,self.ndt.hide)


        else:
            self.ndt.show()
            QTimer.singleShot(2000,self.ndt.hide)


        ptht = "reports/"+cm
        book2 = load_workbook("reports/"+cm+"/"+cm+".xlsx")
        sheet2 = book2.active
        n = 10
        for itm in colB:
            sheet2.cell(row=n,column=2,value=itm)
            n += 1
        n = 10
        for itm in colC:
            sheet2.cell(row=n, column=3, value=itm)
            n += 1
        n = 10
        for itm in colD:
            sheet2.cell(row=n, column=4, value=itm)
            n += 1
        n = 10
        for itm in colE:
            sheet2.cell(row=n, column=5, value=itm)
            n += 1
        book2.save(filename="reports/"+cm+"/"+cm+"_report"+str(len(os.listdir(ptht)))+".xlsx")

    def goback(self):
        widget.setCurrentIndex(1)
        self.comboBox.clear()
        


class AddInfoTakeAttendance(RegisterCourses):
    def __init__(self):
        super(AddInfoTakeAttendance, self).__init__()
        loadUi("uis/addinfotakeattendancescreen.ui", self)
        self.vdstate = False
        self.nocourse.hide()
        self.nofaces.hide()
        date_str = str(date.today())
        # convert str to QDate
        qdate = QtCore.QDate.fromString(date_str, "yyyy-MM-d")
        # Set the format of how the QDate will be displayed in the widget
        self.date.setDisplayFormat("d-MMM-yyyy")
        self.date.setDate(qdate)
        self.okButton.clicked.connect(self.gototakeattendance)
        self.retrieveButton.clicked.connect(self.retrieve)
        self.cancelButton.clicked.connect(self.goback)
        self.pb.setStyleSheet("QProgressBar::chunk{""background-color: rgb(34, 93, 205);""border-radius: 15px;""}")


    def retrieve(self):
        db = mysql.connector.connect(host="localhost", user="jonpol", password="root", database='courses')
        cour = ['TELECOMMUNICATION_ENGINEERING','COMPUTER_ENGINEERING']
        ps = []
        for itm in range(rgc.comboBox4.count()):
            mk = rgc.comboBox4.itemText(itm).split()[0]
            mm = os.listdir("faces/" + mk)
            if len(mm) != 0:
                ps.append(int(len(mm) / 30))
            else:
                ps.append(0)
        for css in cour:
            db1 = mysql.connector.connect(host="localhost", user="jonpol", password="root",
                                          database=css)
            cursor = db.cursor()
            cursor1 = db1.cursor()
            if len(os.listdir("faces")) != 0:
                coursetables = ['TE_COURSES', 'CEIT_COURSES']
                for tables in coursetables:
                    cursor.execute("select * from " + tables)
                    years = [i[0] for i in cursor.description]
                    records = cursor.fetchall()
                    for n in range(1, len(years)):
                        for nm in records:
                            for itm in range(rgc.comboBox4.count()):
                                if (nm[n] == rgc.comboBox4.itemText(itm)):
                                    regnos = []
                                    ptth = os.listdir("faces/" + rgc.comboBox4.itemText(itm).split()[0])
                                    if len(ptth) != 0:
                                        lastid = ptth[-1].split('.')[0]
                                        id = int(lastid) + 1
                                        for m in range(1, int(ptth[-1].split('.')[0]) + 1):
                                            regnos.append(ptth[(m - 1) * 30].split('.')[3])
                                        cursor1.execute("select STUDENT_NAME,REG_NO,IMAGE from " + years[n])
                                        x = 1
                                        for img in cursor1:
                                            count = 0
                                            for fl in regnos:
                                                if img[1] == fl:
                                                    count += 1
                                            if count < 1:
                                                with open("faces/" + nm[n].split()[0] + "/" + str(id) + "." + str(
                                                        x) + "." +
                                                          img[0] + "." + img[1] + ".png", "wb") as file:
                                                    file.write(img[2])
                                                    x += 1
                                                    if x == 31:
                                                        x = 1
                                                        id += 1

                                    else:
                                        id = 1
                                        cursor1.execute("select STUDENT_NAME,REG_NO,IMAGE from " + years[n])
                                        x = 1
                                        for img in cursor1:
                                            with open("faces/" + nm[n].split()[0] + "/" + str(id) + "." + str(x) + "." +
                                                      img[0] + "." + img[1] + ".png", "wb") as file:
                                                file.write(img[2])
                                                x += 1
                                                if x == 31:
                                                    x = 1
                                                    id += 1

                ads = []
                for itm in range(rgc.comboBox4.count()):
                    ptho = rgc.comboBox4.itemText(itm).split()[0]
                    ptth = os.listdir("faces/" + ptho)
                    if len(ptth) != 0:
                        ads.append(int(len(ptth)/30))
                        book = load_workbook("sheets/" + ptho + "/" + ptho + ".xlsx")
                        sheet = book.active
                        n = 1
                        for dir in range(int(ptth[-1].split('.')[0])):
                            sheet.cell(row=n + 9, column=1).value = n
                            sheet.cell(row=n + 9, column=2).value = ptth[dir * 30].split('.')[2]
                            sheet.cell(row=n + 9, column=3).value = ptth[dir * 30].split('.')[3]
                            n += 1
                        book.save(filename="sheets/" + ptho + "/" + ptho + ".xlsx")

                    else:
                        ads.append(0)

                widget = QWidget()
                vbox = QVBoxLayout()
                widget.setLayout(vbox)
                self.scroll.setWidget(widget)
                ls = QListWidget()
                for i in range(len(os.listdir("faces"))):
                    QListWidgetItem(rgc.comboBox4.itemText(i) +
                                    ": " + str(ads[i]) + " Students" +
                                    ", " + str(ads[i] - ps[i]) + " Added""\n", ls)
                    ls.setFont(QFont('Arial', 10))
                    vbox.addWidget(ls)

            else:
                self.nocourse.show()
                QTimer.singleShot(2000,self.nocourse.hide)


            db1.close()


    def imgsandlables(self):
        images = "faces/" + self.comboBox.currentText().split()[0]
        imagePaths = [os.path.join(images, j) for j in os.listdir(images)]
        indfaces = []
        ids = []
        vl = 1
        for imagePath in imagePaths:
            img = Image.open(imagePath).convert('L')  # grayscale
            imgnp = np.array(img, 'uint8')
            id = int(os.path.split(imagePath)[-1].split(".")[0])
            faces = detector.detectMultiScale(imgnp)
            for (x, y, w, h) in faces:
                indfaces.append(imgnp[y:y + h, x:x + w])
                ids.append(id)
                self.pb.setValue(int(vl / len(os.listdir(images)) * 100))
                vl += 1
        return indfaces, ids

    def gototakeattendance(self):
        if self.comboBox.currentText() == "":
            self.nocourse.show()
            QTimer.singleShot(2000,self.nocourse.hide)


        elif len(os.listdir("faces/" + self.comboBox.currentText().split()[0])) == 0:
            self.nofaces.show()
            QTimer.singleShot(2000,self.nofaces.hide)


        elif self.vdstate == True:
            widget.setCurrentIndex(4)

        else:
            faces, ids = self.imgsandlables()
            recognizer.train(faces, np.array(ids))
            names = ['unknown']
            reg = ['unknown']
            m = 1
            imagePaths = os.listdir("faces/" + self.comboBox.currentText().split()[0])
            for n in range(int(imagePaths[-1].split(".")[0])):
                nam = imagePaths[m].split(".")[2]
                names.append(nam)
                rag = imagePaths[m].split(".")[3]
                reg.append(rag)
                m += 30

            widget.setCurrentIndex(4)
            self.pb.setValue(0)

    def goback(self):
        widget.setCurrentIndex(1)
        self.comboBox.clear()
        


class RegisterFaces(QDialog):
    def __init__(self):
        super(RegisterFaces, self).__init__()
        loadUi("uis/registerfacescreen.ui", self)
        self.status = False
        self.stopped.hide()
        self.fieldnotfilled.hide()
        self.first.textChanged.connect(self.auto_cap1)
        self.middle.textChanged.connect(self.auto_cap2)
        self.sir.textChanged.connect(self.auto_cap3)
        self.comboBox1.activated.connect(self.addyears)
        self.comboBox2.activated.connect(self.addcourses)
        self.startButton.clicked.connect(self.registerfaces)
        self.stopButton.clicked.connect(self.stop)
        self.saveButton.clicked.connect(self.upload)
        self.gobackButton.clicked.connect(self.goback)
        self.pb.setStyleSheet("QProgressBar::chunk{""background-color: rgb(34, 93, 205);""border-radius: 15px;""}")
        self.pb2.setStyleSheet("QProgressBar::chunk{""background-color: rgb(34, 93, 205);""border-radius: 15px;""}")


        

    def auto_cap1(self, txt):
        cap_text = txt.title()
        self.first.setText(cap_text)

    def auto_cap2(self, txt):
        cap_text = txt.title()
        self.middle.setText(cap_text)

    def auto_cap3(self, txt):
        cap_text = txt.title()
        self.sir.setText(cap_text)

    def addyears(self):
        db = mysql.connector.connect(host="localhost", user="jonpol", password="root", database='courses')
        cursor = db.cursor()
        self.comboBox2.clear()
        self.comboBox4.clear()
        cursor.execute("select " + self.comboBox1.currentText() + " from FUCULTIES")
        for course in cursor:
            self.comboBox2.addItem(course[0])

    def addcourses(self):
        db = mysql.connector.connect(host="localhost", user="jonpol", password="root", database='courses')
        cursor = db.cursor()
        self.comboBox4.clear()
        cursor.execute("select " + self.comboBox2.currentText() + " from TE_COURSES")
        for course in cursor:
            self.comboBox4.addItem(course[0])

    def upload(self):
        n = 1
        path1 = "img_temp"
        path2 = "img_temp"
        db1 = mysql.connector.connect(host="localhost", user="jonpol", password="root",
                                      database=self.comboBox1.currentText())
        cursor1 = db1.cursor()
        sir_name = self.sir.text()
        first_name = self.first.text()
        middle_name = self.middle.text()
        reg_no = self.reg.text()
        image_path = [os.path.join(path1, i) for i in os.listdir(path2)]
        if self.sir.text() == "" or self.first.text() == "" or self.reg.text() == "":
            self.fieldnotfilled.show()
            QTimer.singleShot(2000,self.fieldnotfilled.hide)

            

        else:
            for img in image_path:
                image = open(img, 'rb').read()
                sql = "insert into " + self.comboBox2.currentText() + " (ID,STUDENT_NAME,REG_NO,COURSES,IMAGE) values(%s,%s,%s,%s,%s)"
                val = (n, sir_name + ", " + first_name + " " + middle_name, reg_no, self.comboBox4.itemText(n-1), image,)
                cursor1.execute(sql, val)
                n += 1
                self.pb2.setValue(int(n * 100 / 30))
            cursor1.close()
            db1.commit()
            for img in image_path:
                os.remove(img)

    def stop(self):
        self.status = True
        time.sleep(1)
        self.t.join()
        self.stopped.show()
        QTimer.singleShot(2000,self.stopped.hide)


    def registerfaces(self):
        rgf.status = False
        self.thread = VideoThread2()
        self.thread.change_pixmap_signal.connect(self.thread.update_image)
        self.t = Thread(target=self.thread.run)
        self.t.start()

    def goback(self):
        widget.setCurrentIndex(1)


class VideoThread2(QThread):

    def run(self):
        scl = [50, 30, 20, 15, 10]
        detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
        count = 1
        while True:
            img_resp = requests.get(url)
            img_arr = np.array(bytearray(img_resp.content), dtype=np.uint8)
            img = cv2.imdecode(img_arr, -1)
            imgs = imutils.resize(img, width=1920, height=1080)
            cv_img = cv2.flip(imgs, 1)

            gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, 1.3, 5)
            for (x, y, w, h) in faces: 

                height, width, channels = cv_img.shape
                cropped = cv_img[y-80:y + h + 80, x-160:x + w + 160]
                cv_img = cv2.resize(cropped, (width, height))
                
                gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
                faces = detector.detectMultiScale(gray, 1.3, 5)
                for (x, y, w, h) in faces:
                    cv2.rectangle(cv_img, (x, y), (x + w, y + h), (255, 255, 255), 2)
                    cv2.imwrite("img_temp/" + str(count) + ".jpg", gray[y:y + h, x:x + w])
                    rgf.pb.setValue(int(count * 100 / 30))
                    if count == 30:
                        rgf.status = True
                        break
                    count += 1
            if True:
                self.change_pixmap_signal.emit(cv_img)

            if rgf.status == True:
                break



    change_pixmap_signal = pyqtSignal(np.ndarray)

    @pyqtSlot(np.ndarray)
    def update_image(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        faces = detector.detectMultiScale(rgb_image, scaleFactor=1.3, minNeighbors=5, )
        for (x, y, w, h) in faces:
            cv2.rectangle(cv_img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(329, 219, Qt.KeepAspectRatio)
        """Updates the image_label with a new opencv image"""
        qt_img = QPixmap.fromImage(p)
        rgf.picture.setPixmap(qt_img)


# main
app = QApplication(sys.argv)
app.setWindowIcon(QIcon("ud"))
widget = QtWidgets.QStackedWidget()
startapp = StartApp()
st = startapp.takeattendance
rgc = startapp.rgstrcourse
vat = startapp.viewattendance
ad = startapp.addinfotakeattendance
rgf = startapp.registerface

try:
    sys.exit(app.exec_())
except:
    print("Exiting")

